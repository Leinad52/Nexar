import re
import unicodedata
from dataclasses import dataclass

from django.db.models import Q
from openpyxl import load_workbook

from .models import Part, Vehicle


@dataclass
class ApplicationImportResult:
    files: int = 0
    rows: int = 0
    linked: int = 0
    skipped: int = 0
    errors: list[str] = None
    preview: list[dict] = None

    def __post_init__(self):
        self.errors = self.errors or []
        self.preview = self.preview or []


PART_HEADERS = {'codigo', 'cod', 'referencia', 'ref', 'sku', 'produto', 'peca', 'codigo peca', 'codigo produto', 'ref peca'}
BRAND_HEADERS = {'marca veiculo', 'montadora', 'fabricante veiculo', 'fabricante', 'marca'}
MODEL_HEADERS = {'modelo', 'veiculo', 'carro', 'aplicacao'}
VERSION_HEADERS = {'versao', 'descricao veiculo', 'descricao modelo'}
YEAR_HEADERS = {'ano', 'anos', 'ano modelo'}
YEAR_START_HEADERS = {'ano inicial', 'inicio', 'de', 'ano de', 'ano ini'}
YEAR_END_HEADERS = {'ano final', 'fim', 'ate', 'ano ate', 'ano fim'}
ENGINE_HEADERS = {'motor', 'motorizacao', 'cilindrada'}
IGNORED_HEADERS = {'linha', 'tipo', 'combustivel', 'obs', 'observacao', 'observacoes'}
APPLY_MARKS = {'x', 'sim', 's', 'ok', 'aplica', 'aplicavel', '1'}
PDF_COLUMN_SPLIT_RE = re.compile(r'\s*\|\s*|\t+|\s{2,}|;')
PDF_KEY_VALUE_RE = re.compile(r'([^:]+):\s*([^:]+?)(?=\s+[A-Za-zÀ-ÿ][^:]{1,30}:|$)')
PDF_FIELD_ALIASES = {
    'part': PART_HEADERS | {'codigo produto', 'produto'},
    'brand': BRAND_HEADERS,
    'model': MODEL_HEADERS,
    'version': VERSION_HEADERS,
    'year': YEAR_HEADERS,
    'year_start': YEAR_START_HEADERS,
    'year_end': YEAR_END_HEADERS,
    'engine': ENGINE_HEADERS,
}
PDF_CANONICAL_HEADERS = {
    'part': 'codigo',
    'brand': 'marca',
    'model': 'modelo',
    'version': 'versao',
    'year': 'ano',
    'year_start': 'ano inicial',
    'year_end': 'ano final',
    'engine': 'motor',
}


def import_application_file(uploaded_file, apply=False, max_preview=40):
    filename = (getattr(uploaded_file, 'name', '') or '').lower()
    result = ApplicationImportResult(files=1)
    if filename.endswith('.pdf'):
        import_application_pdf(uploaded_file, result, apply, max_preview)
        return result

    if not filename.endswith(('.xlsx', '.xlsm')):
        result.errors.append('Formato nao suportado. Envie PDF, XLSX ou XLSM.')
        result.skipped += 1
        return result

    workbook = load_workbook(uploaded_file, read_only=True, data_only=True)
    for sheet in workbook.worksheets:
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            continue

        header_index = find_header_index(rows)
        if header_index is None:
            continue

        headers = [normalize_header(value) for value in rows[header_index]]
        mapping = map_headers(headers)
        for row in rows[header_index + 1:]:
            row_data = {
                headers[index]: normalize_cell(value)
                for index, value in enumerate(row)
                if index < len(headers) and headers[index]
            }
            process_row(row_data, mapping, result, apply, max_preview)

    return result


def import_application_pdf(uploaded_file, result, apply=False, max_preview=40):
    try:
        text = extract_pdf_text(uploaded_file)
    except ImportError:
        result.errors.append('Leitura de PDF indisponivel. Instale as dependencias com: python -m pip install -r requirements.txt')
        result.skipped += 1
        return
    except Exception as error:
        result.errors.append(f'Nao foi possivel ler o PDF: {error}')
        result.skipped += 1
        return

    rows = rows_from_pdf_text(text)
    if not rows:
        result.errors.append('Nenhuma tabela de aplicacao reconhecida no PDF. Se ele for imagem escaneada, sera necessario OCR ou a planilha original.')
        result.skipped += 1
        return

    for row_data, mapping in rows:
        process_row(row_data, mapping, result, apply, max_preview)


def extract_pdf_text(uploaded_file):
    try:
        from pypdf import PdfReader
    except ImportError:
        from PyPDF2 import PdfReader

    if hasattr(uploaded_file, 'seek'):
        uploaded_file.seek(0)

    reader = PdfReader(uploaded_file)
    return '\n'.join(page.extract_text() or '' for page in reader.pages)


def rows_from_pdf_text(text):
    lines = [normalize_pdf_line(line) for line in text.splitlines()]
    lines = [line for line in lines if line]
    return rows_from_pdf_table(lines) or rows_from_pdf_key_values(lines)


def rows_from_pdf_table(lines):
    split_rows = [split_pdf_columns(line) for line in lines]
    header_index = find_header_index(split_rows)
    if header_index is None:
        return []

    headers = [normalize_header(value) for value in split_rows[header_index]]
    mapping = map_headers(headers)
    rows = []
    for columns in split_rows[header_index + 1:]:
        if len([column for column in columns if column]) < 2:
            continue

        row_data = {
            headers[index]: normalize_cell(value)
            for index, value in enumerate(columns)
            if index < len(headers) and headers[index]
        }
        if row_data:
            rows.append((row_data, mapping))
    return rows


def rows_from_pdf_key_values(lines):
    rows = []
    for line in lines:
        row = parse_pdf_key_value_line(line)
        if not row:
            continue

        headers = list(row)
        rows.append((row, map_headers(headers)))
    return rows


def parse_pdf_key_value_line(line):
    row = {}
    for raw_key, raw_value in PDF_KEY_VALUE_RE.findall(line):
        key = normalize_pdf_field(raw_key)
        if key and raw_value.strip():
            row[key] = raw_value.strip()
    return row if row and get_part_codes_from_row(row, map_headers(list(row))) else {}


def normalize_pdf_field(value):
    header = normalize_header(value)
    for canonical, aliases in PDF_FIELD_ALIASES.items():
        if header in {normalize_header(alias) for alias in aliases}:
            return PDF_CANONICAL_HEADERS[canonical]
    return header


def split_pdf_columns(line):
    columns = [column.strip() for column in PDF_COLUMN_SPLIT_RE.split(line)]
    return [column for column in columns if column]


def normalize_pdf_line(line):
    return ' '.join((line or '').replace('\xa0', ' ').split())


def find_header_index(rows):
    for index, row in enumerate(rows[:20]):
        headers = {normalize_header(value) for value in row}
        has_vehicle_columns = headers & (MODEL_HEADERS | BRAND_HEADERS)
        has_application_columns = headers & PART_HEADERS or len([header for header in headers if header]) >= 5
        if has_vehicle_columns and has_application_columns:
            return index
    return None


def map_headers(headers):
    mapped_headers = {
        first_matching_header(headers, PART_HEADERS),
        first_matching_header(headers, BRAND_HEADERS),
        first_matching_header(headers, MODEL_HEADERS),
        first_matching_header(headers, VERSION_HEADERS),
        first_matching_header(headers, YEAR_HEADERS),
        first_matching_header(headers, YEAR_START_HEADERS),
        first_matching_header(headers, YEAR_END_HEADERS),
        first_matching_header(headers, ENGINE_HEADERS),
    }
    mapped_headers.discard('')
    ignored_headers = {normalize_header(header) for header in IGNORED_HEADERS}
    return {
        'part': first_matching_header(headers, PART_HEADERS),
        'brand': first_matching_header(headers, BRAND_HEADERS),
        'model': first_matching_header(headers, MODEL_HEADERS),
        'version': first_matching_header(headers, VERSION_HEADERS),
        'year': first_matching_header(headers, YEAR_HEADERS),
        'year_start': first_matching_header(headers, YEAR_START_HEADERS),
        'year_end': first_matching_header(headers, YEAR_END_HEADERS),
        'engine': first_matching_header(headers, ENGINE_HEADERS),
        'part_columns': [
            header
            for header in headers
            if header and header not in mapped_headers and header not in ignored_headers
        ],
    }


def process_row(row, mapping, result, apply, max_preview):
    part_codes = get_part_codes_from_row(row, mapping)
    if not part_codes:
        result.skipped += 1
        return

    parts = find_parts(part_codes)
    if not parts:
        result.skipped += 1
        result.errors.append(f'Peca nao encontrada: {", ".join(part_codes[:6])}')
        return

    query = build_vehicle_query(row, mapping)
    if not query:
        result.skipped += 1
        return

    vehicles = list(find_vehicles(row, mapping, query)[:300])
    result.rows += 1
    if not vehicles:
        result.skipped += 1
        result.errors.append(f'Nenhum modelo encontrado para: {query}')
        return

    if apply:
        for part in parts:
            part.compatible_vehicles.add(*vehicles)
        result.linked += len(parts) * len(vehicles)

    if len(result.preview) < max_preview:
        result.preview.append(
            {
                'part': ', '.join(str(part) for part in parts),
                'query': query,
                'vehicles': vehicles[:6],
                'vehicle_count': len(vehicles),
            }
        )


def get_part_codes_from_row(row, mapping):
    codes = []
    if mapping['part']:
        codes.extend(split_codes(row.get(mapping['part'], '')))

    for header in mapping['part_columns']:
        value = row.get(header, '')
        if not value:
            continue

        if normalize_header(value) in APPLY_MARKS:
            codes.extend(split_codes(header))
        else:
            codes.extend(split_codes(value))

    return dedupe_preserving_order(codes)


def find_parts(part_codes):
    tokens = part_codes if isinstance(part_codes, list) else split_codes(part_codes)
    query = Q()
    for token in tokens:
        query |= Q(code__iexact=token) | Q(barcode__iexact=token) | Q(notes__icontains=token)
    return list(Part.objects.filter(query)) if query else []


def split_codes(value):
    tokens = []
    for token in re.split(r'[,;/\n\r]+', str(value)):
        token = token.strip().upper()
        if token and normalize_header(token) not in APPLY_MARKS:
            tokens.append(token)
    return tokens


def build_vehicle_query(row, mapping):
    pieces = []
    for key in ('brand', 'model', 'version', 'engine'):
        header = mapping.get(key)
        if header and row.get(header):
            pieces.append(row[header])

    year_start, year_end = get_year_range(row, mapping)
    if year_start and year_start == year_end:
        pieces.append(str(year_start))
    return ' '.join(pieces).strip()


def find_vehicles(row, mapping, query):
    vehicles = search_vehicles_queryset(query)
    year_start, year_end = get_year_range(row, mapping)
    if year_start and year_end:
        ranged = vehicles.filter(Q(year__gte=year_start, year__lte=year_end) | Q(year=0))
        if ranged.exists():
            return ranged
    return vehicles


def get_year_range(row, mapping):
    year_start = parse_year(row.get(mapping['year_start'], '')) if mapping.get('year_start') else None
    year_end = parse_year(row.get(mapping['year_end'], '')) if mapping.get('year_end') else None

    if mapping.get('year') and row.get(mapping['year']):
        years = [int(match) for match in re.findall(r'(?:19|20)\d{2}', row[mapping['year']])]
        if years:
            year_start = year_start or min(years)
            year_end = year_end or max(years)

    if year_start and not year_end:
        year_end = year_start
    if year_end and not year_start:
        year_start = year_end
    return year_start, year_end


def parse_year(value):
    match = re.search(r'(?:19|20)\d{2}', str(value))
    return int(match.group(0)) if match else None


def search_vehicles_queryset(query):
    terms = [term for term in re.split(r'\s+', query) if term]
    vehicles = Vehicle.objects.all()
    for term in terms:
        vehicles = vehicles.filter(
            Q(brand__icontains=term)
            | Q(model__icontains=term)
            | Q(version__icontains=term)
            | Q(year__icontains=term)
            | Q(engine__icontains=term)
            | Q(fuel__icontains=term)
            | Q(fipe_code__icontains=term)
        )
    return vehicles


def first_matching_header(headers, candidates):
    normalized_candidates = {normalize_header(candidate) for candidate in candidates}
    for header in headers:
        if header in normalized_candidates:
            return header
    return ''


def dedupe_preserving_order(values):
    seen = set()
    deduped = []
    for value in values:
        if value not in seen:
            seen.add(value)
            deduped.append(value)
    return deduped


def normalize_cell(value):
    if value is None:
        return ''
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def normalize_header(value):
    value = normalize_cell(value).lower()
    value = unicodedata.normalize('NFKD', value)
    value = ''.join(char for char in value if not unicodedata.combining(char))
    return ' '.join(value.replace('_', ' ').split())
